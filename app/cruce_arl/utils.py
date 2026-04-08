import json
import io
import re
import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST


# ------------------------------------------------------------
# 1. Lectura del archivo Reporte
# ------------------------------------------------------------
def parse_reporte(file_obj) -> pd.DataFrame:
    wb = load_workbook(file_obj, data_only=False)
    ws = wb["Datos"]

    codigo_col_idx = None
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value and str(cell_value).strip() == "Código":
            codigo_col_idx = col
            break

    if codigo_col_idx is None:
        raise Exception("No se encontró la columna 'Código' en la hoja Datos")

    codigo_raw = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=codigo_col_idx)
        val = cell.value
        codigo_raw[row] = val if val is not None else ""

    file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name="Datos", dtype=str, keep_default_na=False)

    if "Código" in df.columns:
        for idx, row in df.iterrows():
            excel_row = idx + 2
            df.at[idx, "Código"] = codigo_raw.get(excel_row, "")

    df["Código"] = df["Código"].replace(['nan', 'None', ''], '')
    return df


# ------------------------------------------------------------
# 2. Lectura del archivo TrabajadoresVigentes
# ------------------------------------------------------------
def _extract_riesgo_from_sheet(ws) -> str:
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if cell.value and str(cell.value).strip().lower().startswith("riesgo"):
                for offset in range(1, 5):
                    neighbour = ws.cell(row=cell.row, column=cell.column + offset)
                    if neighbour.value is not None and str(neighbour.value).strip() not in ("", "nan"):
                        return str(neighbour.value).strip()
    return ""


def parse_trabajadores(file_obj) -> pd.DataFrame:
    xl = pd.ExcelFile(file_obj)
    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True)

    records = []
    for sheet in xl.sheet_names:
        ws = wb[sheet]
        riesgo_val = _extract_riesgo_from_sheet(ws)

        df_sheet = pd.read_excel(xl, sheet_name=sheet, header=None)

        header_row_idx = None
        id_col = None
        for idx, row in df_sheet.iterrows():
            for col_idx, val in enumerate(row):
                if pd.notna(val) and "Identificaci" in str(val):
                    header_row_idx = idx
                    id_col = col_idx
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            continue

        header = df_sheet.iloc[header_row_idx]
        col_map = {}
        for col_idx, val in enumerate(header):
            if pd.notna(val):
                col_map[str(val).strip()] = col_idx

        campos = ["Identificación", "Nombre", "Cargo",
                  "Inicio Vigencia", "EPS", "AFP", "Salario", "Fecha Nac."]

        for data_idx in range(header_row_idx + 1, len(df_sheet)):
            row_data = df_sheet.iloc[data_idx]
            id_val = str(row_data[id_col]) if pd.notna(row_data[id_col]) else ""
            id_val = id_val.strip()
            if not id_val or id_val == "nan" or "NÚMERO" in id_val or id_val == "." or id_val == "0":
                continue

            rec = {"raw_id": id_val, "Riesgo ARL": riesgo_val}
            for field in campos:
                col_idx = col_map.get(field)
                if col_idx is not None:
                    v = row_data[col_idx]
                    rec[field] = str(v).strip() if pd.notna(v) else ""
                else:
                    rec[field] = ""
            records.append(rec)

    if not records:
        raise Exception("No se encontraron datos de trabajadores")

    df = pd.DataFrame(records)

    def split_id(s):
        s = re.sub(r"[^\w\s]", "", s)
        m = re.match(r"^(CC|CE|TI|PA|PEP|PT)\s+(.+)$", s.strip(), re.IGNORECASE)
        if m:
            return m.group(1).upper(), m.group(2).strip()
        digits_only = re.sub(r"[^\d]", "", s)
        if digits_only:
            return "CC", digits_only
        return "CC", s.strip()

    df["Tipo"] = df["raw_id"].apply(lambda x: split_id(x)[0])
    df["ID_Num"] = df["raw_id"].apply(lambda x: split_id(x)[1])
    return df


# ------------------------------------------------------------
# 3. Previsualizaciones para el dashboard
# ------------------------------------------------------------
def build_cruce_preview(df_trab: pd.DataFrame) -> list[dict]:
    preview = []
    for _, row in df_trab.iterrows():
        preview.append({
            "Tipo": row.get("Tipo", ""),
            "Identificación": row.get("ID_Num", ""),
            "Nombre": row.get("Nombre", ""),
            "Cargo": row.get("Cargo", ""),
            "Inicio Vigencia": row.get("Inicio Vigencia", ""),
            "EPS": row.get("EPS", ""),
            "AFP": row.get("AFP", ""),
            "Salario": row.get("Salario", ""),
            "Fecha Nac.": row.get("Fecha Nac.", ""),
            "RIESGO EN ARL": row.get("Riesgo ARL", ""),
            "C. COSTO": "",
            "LIBRA": "",
            "VALIDACION": "",
        })
    return preview


def build_emp_preview(df_rep: pd.DataFrame) -> list[dict]:
    cols_show = [
        "Cédula identificación", "Código", "Apellidos, Nombre",
        "Nombre", "EPS", "AFP", "CCF", "C.COSTO",
        "NIVEL ARL", "Salario Mes",
    ]
    existing = [c for c in cols_show if c in df_rep.columns]
    if not existing:
        existing = list(df_rep.columns[:10])
    return df_rep[existing].head(50).fillna("").to_dict(orient="records")


# ------------------------------------------------------------
# 4. Funciones auxiliares de formato
# ------------------------------------------------------------
def normalize_id(id_val) -> str:
    if pd.isna(id_val):
        return ""
    s = str(id_val).strip()
    return re.sub(r"[^\d]", "", s)


def format_nivel_arl(val) -> str:
    try:
        return f"{int(float(val)):02d}"
    except:
        return ""


def _id_to_int(id_str) -> int | None:
    digits = re.sub(r"[^\d]", "", str(id_str))
    try:
        return int(digits) if digits else None
    except ValueError:
        return None


def format_riesgo_arl(val) -> str:
    try:
        return f"{int(float(val)):02d}"
    except:
        return str(val).strip() if val else ""


def format_codigo_emp(val) -> str:
    if pd.isna(val) or val == "":
        return ""
    try:
        num = float(val)
    except (ValueError, TypeError):
        return str(val).strip()

    if num == int(num) and num >= 10**7:
        int_val = int(num)
        if int_val % 10**7 == 0:
            coef = int_val // 10**7
            return f"{coef:05d}E07"

    if num == int(num):
        return str(int(num))
    return str(num)


# ------------------------------------------------------------
# 4.1 Helpers de formato de celdas
# ------------------------------------------------------------
def _copy_cell_format(src_cell, dst_cell):
    """Copia font, alignment, number_format y border de src a dst."""
    if src_cell.font:
        dst_cell.font = copy.copy(src_cell.font)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format
    if src_cell.border:
        dst_cell.border = copy.copy(src_cell.border)


def _apply_thin_border(cell, left=True, right=True, top=False, bottom=False):
    thin = Side(style='thin')
    cell.border = Border(
        left=thin if left else Side(),
        right=thin if right else Side(),
        top=thin if top else Side(),
        bottom=thin if bottom else Side(),
    )


def _apply_green_fill(cell):
    cell.fill = PatternFill(fill_type="solid", fgColor="92D050")


def _apply_no_fill(cell):
    cell.fill = PatternFill(fill_type=None)


# ------------------------------------------------------------
# 5. Generación del archivo INFRA final
# ------------------------------------------------------------
def generate_infra(infra_bytes: bytes, df_rep: pd.DataFrame, df_trab: pd.DataFrame) -> bytes:
    wb = load_workbook(io.BytesIO(infra_bytes), keep_vba=False)

    if "EMP" not in wb.sheetnames:
        raise Exception("La hoja 'EMP' no existe en INFRA")
    if "Cruce ARL" not in wb.sheetnames:
        raise Exception("La hoja 'Cruce ARL' no existe en INFRA")

    cedula_col = next(
        (col for col in df_rep.columns if "cédula" in col.lower() or "identificacion" in col.lower()),
        "Cédula identificación",
    )
    df_rep["_cedula_norm"] = df_rep[cedula_col].apply(normalize_id)
    rep_dict = {row["_cedula_norm"]: row for _, row in df_rep.iterrows() if row["_cedula_norm"]}

    df_trab["_id_norm"] = df_trab["ID_Num"].apply(normalize_id)
    trab_dict = {row["_id_norm"]: row for _, row in df_trab.iterrows() if row["_id_norm"]}

    _fill_cruce_sheet(wb, df_trab, rep_dict)
    _fill_emp_sheet(wb, df_rep, trab_dict)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ------------------------------------------------------------
# 5.1 Llenar hoja "Cruce ARL"
# ------------------------------------------------------------
def _fill_cruce_sheet(wb, df_trab: pd.DataFrame, rep_dict: dict):
    ws = wb["Cruce ARL"]
    DATA_START = 2
    MAX_DATA_COL = 13

    col_index = {
        "Tipo": 1, "ID_Num": 2, "Nombre": 3, "Cargo": 4,
        "Inicio Vigencia": 5, "EPS": 6, "AFP": 7, "Salario": 8,
        "Fecha Nac.": 9, "Riesgo ARL": 10,
    }

    total = len(df_trab)

    # ── Capturar formatos y fórmulas de referencia UNA sola vez (fuera del loop) ──
    ref_fonts      = {}
    ref_alignments = {}
    ref_numfmts    = {}
    ref_formulas   = {}

    _thin = Side(style='thin')
    _thin_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    for c in range(1, MAX_DATA_COL + 1):
        src = ws.cell(row=2, column=c)
        ref_fonts[c]      = copy.copy(src.font)      if src.font      else None
        ref_alignments[c] = copy.copy(src.alignment) if src.alignment else None
        ref_numfmts[c]    = src.number_format
        if c in (11, 12, 13) and src.value and str(src.value).startswith("="):
            ref_formulas[c] = str(src.value)

    # ── Compilar patrón regex una sola vez (fuera del loop) ──
    _re_row2 = re.compile(r"([A-Z]+)2\b")

    records = df_trab.reset_index(drop=True)

    for i in range(total):
        excel_row = DATA_START + i
        row = records.iloc[i]

        # Aplicar formatos copiados sin llamar copy.copy en cada iteración
        for c in range(1, MAX_DATA_COL + 1):
            cell = ws.cell(row=excel_row, column=c)
            if ref_fonts[c]:      cell.font      = ref_fonts[c]
            if ref_alignments[c]: cell.alignment = ref_alignments[c]
            if ref_numfmts[c]:    cell.number_format = ref_numfmts[c]
            cell.border = _thin_border

        # Escribir valores de datos
        for field, col in col_index.items():
            val = row.get(field, "")
            cell = ws.cell(row=excel_row, column=col)
            if field == "ID_Num":
                cell.value = _id_to_int(val)
            elif field == "Riesgo ARL":
                cell.value = format_riesgo_arl(val) if val else None
                cell.number_format = "@"
            elif field == "Tipo":
                cell.value = val if val else None
                cell.number_format = "@"
            else:
                cell.value = val if val else None

        # Extender fórmulas K, L, M con patrón pre-compilado
        for col_idx, formula in ref_formulas.items():
            cell = ws.cell(row=excel_row, column=col_idx)
            existing = cell.value
            if not existing or not str(existing).startswith("="):
                cell.value = _re_row2.sub(lambda m: f"{m.group(1)}{excel_row}", formula)

    # Limpiar filas sobrantes
    for r in range(DATA_START + total, ws.max_row + 1):
        for c in range(1, MAX_DATA_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    # Formato condicional verde en VALIDACION (col M)
    last_data_row = DATA_START + total - 1
    validacion_range = f"M{DATA_START}:M{last_data_row}"
    green_fill = PatternFill(patternType=None, fgColor="00000000", bgColor="C6EFCE")
    green_font = Font(color="006100")
    dxf_ok = DifferentialStyle(fill=green_fill, font=green_font)
    rule_ok = Rule(type="containsText", operator="containsText", text="OK", dxf=dxf_ok)
    rule_ok.formula = [f'NOT(ISERROR(SEARCH("OK",M{DATA_START})))']
    ws.conditional_formatting._cf_rules.clear()
    ws.conditional_formatting.add(validacion_range, rule_ok)

    print(f"Cruce ARL actualizado: {total} filas, CF verde en {validacion_range}")


# ------------------------------------------------------------
# 5.2 Llenar hoja "EMP"
# ------------------------------------------------------------
def _fill_emp_sheet(wb, df_rep: pd.DataFrame, trab_dict: dict):
    from openpyxl.utils import get_column_letter
    ws = wb["EMP"]
    HEADER_ROW = 1
    DATA_START = 2

    col_c_costo    = None
    col_validacion = None
    col_name_to_idx = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=HEADER_ROW, column=c).value
        if val:
            col_name_to_idx[str(val).strip()] = c
            if str(val).strip() == "C.COSTO":
                col_c_costo = c
            if str(val).strip() == "VALIDACION":
                col_validacion = c

    MAX_DATA_COL = max(col_name_to_idx.values()) if col_name_to_idx else 140

    # ── Capturar formatos y fórmulas de referencia UNA sola vez (fuera del loop) ──
    ref_fonts      = {}
    ref_alignments = {}
    ref_numfmts    = {}
    ref_borders_lr = {}
    formula_cols: dict[int, str] = {}

    for c in range(1, MAX_DATA_COL + 1):
        src = ws.cell(row=DATA_START, column=c)
        ref_fonts[c]      = copy.copy(src.font)      if src.font      else None
        ref_alignments[c] = copy.copy(src.alignment) if src.alignment else None
        ref_numfmts[c]    = src.number_format

        b = src.border
        ref_borders_lr[c] = Border(
            left=copy.copy(b.left)  if b.left.style  else Side(),
            right=copy.copy(b.right) if b.right.style else Side(),
            top=Side(),
            bottom=Side(),
        )

        if src.value and str(src.value).startswith("="):
            formula_cols[c] = str(src.value)

    # Mapeo reporte → columna EMP calculado UNA sola vez (fuera del loop)
    reporte_to_emp: dict[str, int] = {}
    for col in df_rep.columns:
        col_norm = col.strip()
        if col_norm in col_name_to_idx:
            reporte_to_emp[col] = col_name_to_idx[col_norm]
        else:
            for emp_col, idx in col_name_to_idx.items():
                if emp_col.strip().lower() == col_norm.lower():
                    reporte_to_emp[col] = idx
                    break

    skip_cols = set(formula_cols.keys()) | {94}

    # ── Compilar patrón regex una sola vez (fuera del loop) ──
    _re_row2 = re.compile(r"([A-Z]+)2\b")

    n_rows = len(df_rep)
    records = df_rep.reset_index(drop=True)

    for i in range(n_rows):
        excel_row = DATA_START + i
        rep_row = records.iloc[i]

        # Aplicar formatos (solo para filas nuevas, no la de referencia)
        if excel_row != DATA_START:
            for c in range(1, MAX_DATA_COL + 1):
                cell = ws.cell(row=excel_row, column=c)
                if ref_fonts[c]:      cell.font      = ref_fonts[c]
                if ref_alignments[c]: cell.alignment = ref_alignments[c]
                if ref_numfmts[c]:    cell.number_format = ref_numfmts[c]
                cell.border = ref_borders_lr[c]

        # Escribir datos del reporte
        for rep_col, col_idx in reporte_to_emp.items():
            if col_idx in skip_cols:
                continue
            cell = ws.cell(row=excel_row, column=col_idx)
            if cell.value and str(cell.value).startswith("="):
                continue
            val = rep_row.get(rep_col)
            cell.value = val if not pd.isna(val) else None

        # C.COSTO desde "CCF"
        if col_c_costo is not None:
            ccf_val = rep_row.get("CCF", "")
            if ccf_val and str(ccf_val) not in ("nan", "None", ""):
                cell = ws.cell(row=excel_row, column=col_c_costo)
                cell.value = str(ccf_val)
                cell.number_format = "@"

        # Código (columna B)
        codigo_val = rep_row.get("Código", "")
        if codigo_val and str(codigo_val) not in ("nan", "None", ""):
            cell = ws.cell(row=excel_row, column=2)
            cell.value = format_codigo_emp(codigo_val)
            cell.number_format = "@"

        # NIVEL ARL formateado
        nivel_arl = format_nivel_arl(rep_row.get("NIVEL ARL", ""))
        ws.cell(row=excel_row, column=94).value = nivel_arl

        # Propagar fórmulas con patrón pre-compilado
        for col_idx, ref_formula in formula_cols.items():
            cell = ws.cell(row=excel_row, column=col_idx)
            if not cell.value or not str(cell.value).startswith("="):
                cell.value = _re_row2.sub(lambda m: f"{m.group(1)}{excel_row}", ref_formula)

    # Formato condicional verde en VALIDACION
    if col_validacion:
        last_data_row = DATA_START + n_rows - 1
        val_col_letter = get_column_letter(col_validacion)
        val_range = f"{val_col_letter}{DATA_START}:{val_col_letter}{last_data_row}"
        green_fill = PatternFill(patternType=None, fgColor="00000000", bgColor="C6EFCE")
        green_font = Font(color="006100")
        dxf_ok = DifferentialStyle(fill=green_fill, font=green_font)
        rule_ok = Rule(type="containsText", operator="containsText", text="OK", dxf=dxf_ok)
        rule_ok.formula = [f'NOT(ISERROR(SEARCH("OK",{val_col_letter}{DATA_START})))']
        ws.conditional_formatting._cf_rules.clear()
        ws.conditional_formatting.add(val_range, rule_ok)
        print(f"EMP actualizada: {n_rows} filas, CF verde en {val_range}")
    else:
        print(f"EMP actualizada: {n_rows} filas (columna VALIDACION no encontrada)")


# ------------------------------------------------------------
# 6. Vistas de Django  (sin cambios)
# ------------------------------------------------------------
def index(request):
    if request.method == "GET":
        return render(request, "upload.html")

    reporte_file = request.FILES.get("reporte")
    trabajadores_file = request.FILES.get("trabajadores")
    infra_file = request.FILES.get("infra")

    errors = []
    if not reporte_file:
        errors.append("Falta el archivo Reporte de empleados.")
    if not trabajadores_file:
        errors.append("Falta el archivo Trabajadores Vigentes.")
    if not infra_file:
        errors.append("Falta el archivo INFRA base.")

    if errors:
        return render(request, "upload.html", {"errors": errors})

    try:
        df_rep = parse_reporte(reporte_file)
    except Exception as e:
        return render(request, "upload.html", {"errors": [f"Error leyendo Reporte: {e}"]})

    try:
        df_trab = parse_trabajadores(trabajadores_file)
    except Exception as e:
        return render(request, "upload.html", {"errors": [f"Error leyendo Trabajadores: {e}"]})

    infra_bytes = infra_file.read()
    request.session["infra_bytes"] = list(infra_bytes)
    request.session["reporte_json"] = df_rep.to_json(orient="records", date_format="iso", default_handler=str)
    request.session["trabajadores_json"] = df_trab.to_json(orient="records", date_format="iso", default_handler=str)

    cruce_preview = build_cruce_preview(df_trab)
    emp_preview = build_emp_preview(df_rep)

    context = {
        "cruce_preview": cruce_preview[:20],
        "emp_preview": emp_preview[:20],
        "total_trabajadores": len(df_trab),
        "total_empleados": len(df_rep),
        "cruce_cols": list(cruce_preview[0].keys()) if cruce_preview else [],
        "emp_cols": list(emp_preview[0].keys()) if emp_preview else [],
    }
    return render(request, "dashboard.html", context)


@require_POST
def download_infra(request):
    infra_bytes_list = request.session.get("infra_bytes")
    reporte_json = request.session.get("reporte_json")
    trabajadores_json = request.session.get("trabajadores_json")

    if not infra_bytes_list or not reporte_json or not trabajadores_json:
        return JsonResponse({"error": "Sesión expirada. Vuelve a cargar los archivos."}, status=400)

    infra_bytes = bytes(infra_bytes_list)
    df_rep = pd.read_json(io.StringIO(reporte_json), orient="records")
    df_trab = pd.read_json(io.StringIO(trabajadores_json), orient="records")

    try:
        output_bytes = generate_infra(infra_bytes, df_rep, df_trab)
    except Exception as e:
        return JsonResponse({"error": f"Error generando archivo: {e}"}, status=500)

    response = HttpResponse(output_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="INFRA_Cruce_ARL_actualizado.xlsx"'
    return response